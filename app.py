from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash
from email_service import send_budget_email, send_test_email

# Load environment variables from .env file if it exists
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, that's okay

app = Flask(__name__)
# Use environment variable for secret key, or generate one
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(24).hex())

# Support both SQLite (local) and PostgreSQL (production)
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Heroku/Production: DATABASE_URL format is postgresql://...
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///expenses.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def get_current_user_id():
    """Get the current logged-in user ID"""
    return session.get('user_id')


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'email': self.email,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }


class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.String(20), nullable=False)
    category = db.Column(db.String(100), nullable=False)
    subcategory = db.Column(db.String(100), nullable=True)  # Optional subcategory for detailed tracking
    description = db.Column(db.String(500), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    is_recurring = db.Column(db.Boolean, default=False, nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)  # For tracking active/cancelled subscriptions
    is_bill = db.Column(db.Boolean, default=False, nullable=False)  # For subscriptions that are bills (chequing)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref=db.backref('expenses', lazy=True))

    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date,
            'category': self.category,
            'subcategory': self.subcategory,
            'description': self.description,
            'amount': self.amount,
            'is_recurring': self.is_recurring,
            'is_active': self.is_active,
            'is_bill': self.is_bill,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }


class BudgetLimit(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    month = db.Column(db.String(7), nullable=False)  # Format: YYYY-MM
    fixed_bills_loans = db.Column(db.Float, default=600, nullable=False)
    variable_spending = db.Column(db.Float, default=800, nullable=False)
    investing_min = db.Column(db.Float, default=1500, nullable=False)
    investing_max = db.Column(db.Float, default=1800, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    user = db.relationship('User', backref=db.backref('budget_limits', lazy=True))

    def to_dict(self):
        return {
            'id': self.id,
            'month': self.month,
            'fixed_bills_loans': self.fixed_bills_loans,
            'variable_spending': self.variable_spending,
            'investing_min': self.investing_min,
            'investing_max': self.investing_max
        }


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username_or_email = request.form.get('username')
        password = request.form.get('password')
        
        if not username_or_email or not password:
            return render_template('login.html', error='Username/email and password required')
        
        # Try to find user by username or email
        user = User.query.filter(
            (User.username == username_or_email) | (User.email == username_or_email)
        ).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session.permanent = True
            return redirect(url_for('index'))
        
        return render_template('login.html', error='Invalid username/email or password')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        # Validation
        if not username or not email or not password:
            return render_template('register.html', error='All fields are required')
        
        if password != confirm_password:
            return render_template('register.html', error='Passwords do not match')
        
        if len(password) < 6:
            return render_template('register.html', error='Password must be at least 6 characters')
        
        # Check if username or email already exists
        if User.query.filter_by(username=username).first():
            return render_template('register.html', error='Username already taken')
        
        if User.query.filter_by(email=email).first():
            return render_template('register.html', error='Email already registered')
        
        # Create new user
        user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password)
        )
        db.session.add(user)
        db.session.commit()
        
        # Auto-login after registration
        session['user_id'] = user.id
        session['username'] = user.username
        session.permanent = True
        
        return redirect(url_for('index'))
    
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return render_template('index.html')


@app.route('/api/expenses', methods=['GET'])
@login_required
def get_expenses():
    user_id = get_current_user_id()
    month = request.args.get('month')  # Format: YYYY-MM
    year = request.args.get('year')
    
    expenses = Expense.query.filter_by(user_id=user_id)
    
    if month and year:
        # Filter by specific month and year
        expenses = expenses.filter(
            Expense.date.like(f'{year}-{month}%')
        )
    elif month:
        # Filter by month in current year
        current_year = datetime.now().year
        expenses = expenses.filter(
            Expense.date.like(f'{current_year}-{month}%')
        )
    
    expenses = expenses.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    return jsonify([expense.to_dict() for expense in expenses])


@app.route('/api/expenses/all', methods=['GET'])
@login_required
def get_all_expenses():
    """Get all expenses for charts and analysis"""
    user_id = get_current_user_id()
    expenses = Expense.query.filter_by(user_id=user_id).order_by(Expense.date.asc()).all()
    return jsonify([expense.to_dict() for expense in expenses])


@app.route('/api/expenses/generate-recurring', methods=['POST'])
@login_required
def generate_recurring_expenses():
    """Auto-generate recurring expenses for a given month"""
    data = request.json
    target_month = data.get('month')  # Format: YYYY-MM
    
    if not target_month:
        return jsonify({'error': 'Month parameter required'}), 400
    
    # Get all active recurring expenses for current user
    user_id = get_current_user_id()
    recurring_expenses = Expense.query.filter_by(
        user_id=user_id,
        is_recurring=True,
        is_active=True
    ).all()
    
    generated_count = 0
    
    for recurring in recurring_expenses:
        # Check if expense already exists for this month
        existing = Expense.query.filter(
            Expense.user_id == user_id,
            Expense.description == recurring.description,
            Expense.category == recurring.category,
            Expense.amount == recurring.amount,
            Expense.date.like(f'{target_month}%')
        ).first()
        
        if not existing:
            # Generate new expense for this month
            # Use first day of the month
            new_date = f"{target_month}-01"
            
            new_expense = Expense(
                user_id=user_id,
                date=new_date,
                category=recurring.category,
                description=recurring.description,
                amount=recurring.amount,
                is_recurring=True,
                is_active=True,
                is_bill=recurring.is_bill
            )
            db.session.add(new_expense)
            generated_count += 1
    
    db.session.commit()
    return jsonify({'generated': generated_count})


@app.route('/api/expenses', methods=['POST'])
@login_required
def add_expense():
    user_id = get_current_user_id()
    data = request.json
    category = data['category']
    is_recurring = data.get('is_recurring', False)
    # Subscriptions are active by default
    is_active = True if category == 'Subscription' else data.get('is_active', True)
    # is_bill is only relevant for subscriptions
    is_bill = data.get('is_bill', False) if category == 'Subscription' else False
    
    expense = Expense(
        user_id=user_id,
        date=data['date'],
        category=category,
        subcategory=data.get('subcategory'),  # Optional subcategory
        description=data['description'],
        amount=float(data['amount']),
        is_recurring=is_recurring,
        is_active=is_active,
        is_bill=is_bill
    )
    db.session.add(expense)
    db.session.commit()
    return jsonify(expense.to_dict()), 201


@app.route('/api/expenses/<int:expense_id>', methods=['PUT'])
@login_required
def update_expense(expense_id):
    expense = Expense.query.get_or_404(expense_id)
    data = request.json
    
    expense.date = data['date']
    expense.category = data['category']
    expense.subcategory = data.get('subcategory')  # Optional subcategory
    expense.description = data['description']
    expense.amount = float(data['amount'])
    expense.is_recurring = data.get('is_recurring', False)
    expense.is_active = data.get('is_active', True)
    # is_bill is only relevant for subscriptions
    if data['category'] == 'Subscription':
        expense.is_bill = data.get('is_bill', False)
    else:
        expense.is_bill = False
    
    db.session.commit()
    return jsonify(expense.to_dict()), 200


@app.route('/api/expenses/<int:expense_id>/cancel', methods=['POST'])
@login_required
def cancel_subscription(expense_id):
    user_id = get_current_user_id()
    expense = Expense.query.filter_by(id=expense_id, user_id=user_id).first_or_404()
    if expense.category != 'Subscription':
        return jsonify({'error': 'Only subscriptions can be cancelled'}), 400
    
    expense.is_active = False
    db.session.commit()
    return jsonify(expense.to_dict()), 200


@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
@login_required
def delete_expense(expense_id):
    user_id = get_current_user_id()
    expense = Expense.query.filter_by(id=expense_id, user_id=user_id).first_or_404()
    db.session.delete(expense)
    db.session.commit()
    return jsonify({'message': 'Expense deleted successfully'}), 200


@app.route('/api/export/csv')
@login_required
def export_csv():
    user_id = get_current_user_id()
    month = request.args.get('month')
    year = request.args.get('year')
    
    expenses = Expense.query.filter_by(user_id=user_id)
    
    if month and year:
        expenses = expenses.filter(Expense.date.like(f'{year}-{month}%'))
    
    expenses = expenses.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['Date', 'Category', 'Description', 'Amount', 'Recurring', 'Active', 'Is Bill'])
    
    # Write data
    for expense in expenses:
        writer.writerow([expense.date, expense.category, expense.description, expense.amount, 'Yes' if expense.is_recurring else 'No', 'Yes' if expense.is_active else 'No', 'Yes' if expense.is_bill else 'No'])
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'expenses_{datetime.now().strftime("%Y%m%d")}.csv'
    )


@app.route('/api/budget-limits', methods=['GET'])
@login_required
def get_budget_limits():
    month = request.args.get('month')  # Format: YYYY-MM
    if not month:
        return jsonify({'error': 'Month parameter required'}), 400
    
    budget_limit = BudgetLimit.query.filter_by(month=month).first()
    if budget_limit:
        return jsonify(budget_limit.to_dict())
    else:
        # Return defaults if no budget limit exists for this month
        return jsonify({
            'month': month,
            'fixed_bills_loans': 600,
            'variable_spending': 800,
            'investing_min': 1500,
            'investing_max': 1800
        })


@app.route('/api/budget-limits', methods=['POST'])
@login_required
def save_budget_limits():
    user_id = get_current_user_id()
    data = request.json
    month = data.get('month')
    
    if not month:
        return jsonify({'error': 'Month parameter required'}), 400
    
    budget_limit = BudgetLimit.query.filter_by(user_id=user_id, month=month).first()
    
    if budget_limit:
        # Update existing
        budget_limit.fixed_bills_loans = float(data.get('fixed_bills_loans', 600))
        budget_limit.variable_spending = float(data.get('variable_spending', 800))
        budget_limit.investing_min = float(data.get('investing_min', 1500))
        budget_limit.investing_max = float(data.get('investing_max', 1800))
        budget_limit.updated_at = datetime.utcnow()
    else:
        # Create new
        budget_limit = BudgetLimit(
            user_id=user_id,
            month=month,
            fixed_bills_loans=float(data.get('fixed_bills_loans', 600)),
            variable_spending=float(data.get('variable_spending', 800)),
            investing_min=float(data.get('investing_min', 1500)),
            investing_max=float(data.get('investing_max', 1800))
        )
        db.session.add(budget_limit)
    
    db.session.commit()
    return jsonify(budget_limit.to_dict()), 200


@app.route('/api/export/excel')
@login_required
def export_excel():
    user_id = get_current_user_id()
    month = request.args.get('month')
    year = request.args.get('year')
    
    expenses = Expense.query.filter_by(user_id=user_id)
    
    if month and year:
        expenses = expenses.filter(Expense.date.like(f'{year}-{month}%'))
    
    expenses = expenses.order_by(Expense.date.desc(), Expense.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Header row
    headers = ['Date', 'Category', 'Description', 'Amount', 'Recurring', 'Active', 'Is Bill']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, expense in enumerate(expenses, 2):
        ws.cell(row=row_num, column=1, value=expense.date)
        ws.cell(row=row_num, column=2, value=expense.category)
        ws.cell(row=row_num, column=3, value=expense.description)
        ws.cell(row=row_num, column=4, value=expense.amount)
        ws.cell(row=row_num, column=5, value='Yes' if expense.is_recurring else 'No')
        ws.cell(row=row_num, column=6, value='Yes' if expense.is_active else 'No')
        ws.cell(row=row_num, column=7, value='Yes' if expense.is_bill else 'No')
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'expenses_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/api/send-budget-email', methods=['POST'])
@login_required
def send_budget_email_api():
    """Send weekly budget report email"""
    try:
        data = request.json
        recipient_email = data.get('email')
        month = data.get('month')  # Format: YYYY-MM
        
        if not recipient_email:
            return jsonify({'error': 'Email address required'}), 400
        
        # Get expenses for the month
        user_id = get_current_user_id()
        if not month:
            month = datetime.now().strftime('%Y-%m')
        
        try:
            year, month_num = month.split('-')
            year = int(year)
            month_num = int(month_num)
        except (ValueError, AttributeError):
            # Fallback to current month if format is wrong
            now = datetime.now()
            year = now.year
            month_num = now.month
            month = f'{year}-{month_num:02d}'
        
        expenses = Expense.query.filter(
            Expense.user_id == user_id,
            Expense.date.like(f'{year}-{month_num:02d}%')
        ).all()
        
        # Calculate totals
        fixed_bills_loans_spent = sum(
            exp.amount for exp in expenses 
            if exp.category in ['Bills', 'Loans'] or (exp.category == 'Subscription' and exp.is_bill)
        )
        
        # Variable spending: everything except Bills, Loans, Income, Investment, Payment, and Subscription bills
        variable_spending_spent = sum(
            exp.amount for exp in expenses
            if exp.category not in ['Bills', 'Loans', 'Income', 'Investment', 'Payment'] 
            and not (exp.category == 'Subscription' and exp.is_bill)
        )
        
        investment_total = sum(exp.amount for exp in expenses if exp.category == 'Investment')
        income_total = sum(exp.amount for exp in expenses if exp.category == 'Income')
        
        # Get budget limits
        budget_limit = BudgetLimit.query.filter_by(user_id=user_id, month=month).first()
        if not budget_limit:
            # Use defaults
            fixed_bills_loans_limit = 600
            variable_spending_limit = 800
            investment_min = 1500
            investment_max = 1800
        else:
            fixed_bills_loans_limit = budget_limit.fixed_bills_loans
            variable_spending_limit = budget_limit.variable_spending
            investment_min = budget_limit.investing_min
            investment_max = budget_limit.investing_max
        
        # Calculate remaining buffer
        remaining_buffer = income_total - (fixed_bills_loans_spent + variable_spending_spent + investment_total)
        
        # Get top categories
        category_totals = {}
        for exp in expenses:
            if exp.category not in ['Income', 'Investment', 'Payment']:
                category_totals[exp.category] = category_totals.get(exp.category, 0) + exp.amount
        
        top_categories = [
            {'category': cat, 'total': total}
            for cat, total in sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:5]
        ]
        
        # Format month display
        month_display = datetime(year, month_num, 1).strftime('%B %Y')
        
        # Prepare budget data
        budget_data = {
            'month': month_display,
            'fixed_bills_loans_spent': fixed_bills_loans_spent,
            'fixed_bills_loans_limit': fixed_bills_loans_limit,
            'variable_spending_spent': variable_spending_spent,
            'variable_spending_limit': variable_spending_limit,
            'investment_total': investment_total,
            'investment_min': investment_min,
            'investment_max': investment_max,
            'income_total': income_total,
            'remaining_buffer': remaining_buffer,
            'top_categories': top_categories
        }
        
        # Send email
        try:
            send_budget_email(recipient_email, budget_data)
        except Exception as email_error:
            print(f"Email sending error: {email_error}")
            return jsonify({'error': f'Failed to send email: {str(email_error)}'}), 500
        
        return jsonify({'success': True, 'message': 'Budget report sent successfully!'})
        
    except ValueError as e:
        print(f"ValueError in send_budget_email_api: {e}")
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        print(f"Exception in send_budget_email_api: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to send email: {str(e)}'}), 500


@app.route('/api/send-test-email', methods=['POST'])
@login_required
def send_test_email_api():
    """Send a test email to verify email configuration"""
    try:
        data = request.json
        recipient_email = data.get('email')
        
        if not recipient_email:
            return jsonify({'error': 'Email address required'}), 400
        
        send_test_email(recipient_email)
        
        return jsonify({'success': True, 'message': 'Test email sent successfully!'})
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'Failed to send email: {str(e)}'}), 500


def migrate_database():
    """Add is_recurring, is_active, is_bill, and user_id columns if they don't exist"""
    try:
        # Check if columns exist by querying table info
        with db.engine.begin() as conn:
            # Check expense table
            result = conn.execute(db.text("PRAGMA table_info(expense)"))
            columns = [row[1] for row in result.fetchall()]
            
            if 'is_recurring' not in columns:
                # Column doesn't exist, add it
                conn.execute(db.text('ALTER TABLE expense ADD COLUMN is_recurring BOOLEAN DEFAULT 0'))
                print("✓ Added is_recurring column to expense table")
            
            if 'is_active' not in columns:
                # Column doesn't exist, add it
                conn.execute(db.text('ALTER TABLE expense ADD COLUMN is_active BOOLEAN DEFAULT 1'))
                print("✓ Added is_active column to expense table")
            
            if 'is_bill' not in columns:
                # Column doesn't exist, add it
                conn.execute(db.text('ALTER TABLE expense ADD COLUMN is_bill BOOLEAN DEFAULT 0'))
                print("✓ Added is_bill column to expense table")
            
            if 'user_id' not in columns:
                # Column doesn't exist, add it
                # First, create a default user if user table exists
                try:
                    user_result = conn.execute(db.text("SELECT id FROM user LIMIT 1"))
                    default_user_id = user_result.fetchone()
                    if default_user_id:
                        default_user_id = default_user_id[0]
                    else:
                        default_user_id = 1  # Will be created when User table is created
                except:
                    default_user_id = 1
                
                conn.execute(db.text(f'ALTER TABLE expense ADD COLUMN user_id INTEGER DEFAULT {default_user_id}'))
                print("✓ Added user_id column to expense table")
            
            if 'subcategory' not in columns:
                # Column doesn't exist, add it
                conn.execute(db.text('ALTER TABLE expense ADD COLUMN subcategory VARCHAR(100)'))
                print("✓ Added subcategory column to expense table")
            
            # Check budget_limit table
            try:
                result = conn.execute(db.text("PRAGMA table_info(budget_limit)"))
                budget_columns = [row[1] for row in result.fetchall()]
                
                if 'user_id' not in budget_columns:
                    try:
                        user_result = conn.execute(db.text("SELECT id FROM user LIMIT 1"))
                        default_user_id = user_result.fetchone()
                        if default_user_id:
                            default_user_id = default_user_id[0]
                        else:
                            default_user_id = 1
                    except:
                        default_user_id = 1
                    
                    conn.execute(db.text(f'ALTER TABLE budget_limit ADD COLUMN user_id INTEGER DEFAULT {default_user_id}'))
                    print("✓ Added user_id column to budget_limit table")
            except Exception as e:
                print(f"Budget limit table check: {e}")
            
            print("✓ Database schema migration complete")
            
            # Create budget_limit table if it doesn't exist
            try:
                result = conn.execute(db.text("SELECT name FROM sqlite_master WHERE type='table' AND name='budget_limit'"))
                if not result.fetchone():
                    conn.execute(db.text("""
                        CREATE TABLE budget_limit (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            month VARCHAR(7) NOT NULL,
                            fixed_bills_loans FLOAT NOT NULL DEFAULT 600,
                            variable_spending FLOAT NOT NULL DEFAULT 800,
                            investing_min FLOAT NOT NULL DEFAULT 1500,
                            investing_max FLOAT NOT NULL DEFAULT 1800,
                            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                        )
                    """))
                    print("✓ Created budget_limit table")
            except Exception as e:
                print(f"Budget limit table check: {e}")
    except Exception as e:
        print(f"Migration check: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        migrate_database()
    app.run(debug=True, host='127.0.0.1', port=5000)

